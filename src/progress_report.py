import os
import datetime
from task_management import TaskManagement

DASHBOARD_PATH = os.path.join('docs', 'project_management', 'detailed_progress_dashboard.md')
IMPORTANCE_URGENCY_REPORT_PATH = os.path.join('docs', 'project_management', 'importance_urgency_report.md')

def generate_report(tm: TaskManagement):
    tm.calculate_urgency_importance()
    tasks = tm.prioritize_tasks()

    completed = []
    in_progress = []
    pending = []

    for task in tasks:
        status = task.status.lower()
        line = "- {} (Urgency: {:.2f}, Importance: {:.2f})".format(task.title, task.urgency, task.importance)
        if status == 'completed':
            completed.append(line)
        elif status == 'in_progress':
            in_progress.append(line)
        else:
            pending.append(line)

    report = "# Progress Report - {}\n\n".format(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    report += "## Completed Activities\n"
    report += "\n".join(completed) if completed else "- None"
    report += "\n\n## In-Progress Activities\n"
    report += "\n".join(in_progress) if in_progress else "- None"
    report += "\n\n## Pending Activities\n"
    report += "\n".join(pending) if pending else "- None"
    report += "\n\n*This report is generated automatically and reflects real-time project status including urgency and importance metrics.*\n"

    with open(DASHBOARD_PATH, 'w', encoding='utf-8') as f:
        f.write(report)
    print(f"Progress report updated at {DASHBOARD_PATH}")

def generate_progress_dashboard_report(tm: TaskManagement):
    """
    Generate the progress dashboard markdown file at docs/project_management/progress_dashboard.md
    with two-level progress summary, urgent tasks by level, and Eisenhower matrix.
    """
    import os
    import re

    tm.calculate_urgency_importance()
    tasks = tm.prioritize_tasks()

    # Define all 11 phases with descriptions
    phase_descriptions = {
        1: "Setup and Initialization",
        2: "GitHub Integration",
        3: "Task Management",
        4: "Documentation and Reporting",
        5: "Communication and Feedback",
        6: "Automation and Extensibility",
        7: "Security and Permissions",
        8: "Usability and CLI",
        9: "Backup and Recovery",
        10: "Standards Compliance and Multi-Method Support",
        11: "Final Testing, Deployment, and Maintenance",
        None: "Unassigned or Initial phase"
    }

    # Build hierarchical structure: phases and their subtasks
    phases = {phase: {"description": desc, "tasks": []} for phase, desc in phase_descriptions.items()}

    # Assign tasks to phases based on parent_id or None
    for task in tasks:
        phase_id = task.parent_id if task.parent_id in phase_descriptions else None
        phases[phase_id]["tasks"].append(task)

    # Function to calculate progress for a list of tasks
    def calculate_progress(task_list):
        total = len(task_list)
        if total == 0:
            return 0, 0, 0  # completed, total, percent
        completed = sum(1 for t in task_list if t.status.lower() == "completed")
        # If workflow_progress_percentage method exists, average it
        progress_sum = 0
        count_with_progress = 0
        for t in task_list:
            prog = 0
            if hasattr(t, "workflow_progress_percentage"):
                if callable(t.workflow_progress_percentage):
                    prog = t.workflow_progress_percentage()
                else:
                    prog = t.workflow_progress_percentage
            if prog > 0:
                progress_sum += prog
                count_with_progress += 1
        avg_progress = (progress_sum / count_with_progress) if count_with_progress > 0 else (completed / total) * 100
        return completed, total, avg_progress

    # Build markdown content
    md = "# Project Management Dashboard\n\n"
    md += "## Current Progress Summary\n\n"
    md += "| Phase | Description | Completed Tasks | Total Tasks | Progress (%) |\n"
    md += "|-------|-------------|-----------------|-------------|--------------|\n"

    # Add phases and subtasks with indentation
    for phase_id in sorted(phase_descriptions.keys(), key=lambda x: (x is None, x)):
        phase = phases[phase_id]
        completed, total, progress_percent = calculate_progress(phase["tasks"])
        phase_name = f"Phase {phase_id}" if phase_id is not None else "Unassigned"
        md += f"| **{phase_name}** | {phase['description']} | {completed} | {total} | {progress_percent:.0f}% |\n"

        # Group subtasks by their parent task id if any (simulate two-level)
        subtasks_map = {}
        for t in phase["tasks"]:
            parent = getattr(t, "parent_task_id", None)
            if parent not in subtasks_map:
                subtasks_map[parent] = []
            subtasks_map[parent].append(t)

        # Show subtasks under phase (those with parent_task_id == None are top-level tasks)
        for parent_id, subtasks in subtasks_map.items():
            if parent_id is None:
                continue  # skip top-level tasks here
            # Calculate progress for subtasks group
            c, tot, p = calculate_progress(subtasks)
            # Find parent task title for display
            parent_title = next((t.title for t in phase["tasks"] if t.id == parent_id), "Subtasks")
            md += f"| &nbsp;&nbsp;&nbsp;{parent_title} | Subtasks | {c} | {tot} | {p:.0f}% |\n"

    # Task Details with Urgency and Importance
    md += "\n## Task Details with Urgency and Importance\n\n"
    md += "| Task ID | Title | Urgency | Importance | Status |\n"
    md += "|---------|-------|---------|------------|--------|\n"
    for task in tasks:
        md += f"| {task.id} | {task.title} | {task.urgency:.2f} | {task.importance:.2f} | {task.status} |\n"

    # Urgent Tasks grouped by hierarchical levels
    md += "\n## Urgent Tasks by Hierarchical Levels\n\n"
    # Group tasks by level extracted from title e.g. "Level 1.1"
    level_map = {}
    level_pattern = re.compile(r"Level ([\\d\\.]+)")
    for task in tasks:
        match = level_pattern.search(task.title)
        level = match.group(1) if match else "1"
        if level not in level_map:
            level_map[level] = []
        level_map[level].append(task)

    for level in sorted(level_map.keys()):
        md += f"### Level {level}\n\n"
        md += "| Task ID | Title | Urgency | Importance | Status |\n"
        md += "|---------|-------|---------|------------|--------|\n"
        for task in level_map[level]:
            md += f"| {task.id} | {task.title} | {task.urgency:.2f} | {task.importance:.2f} | {task.status} |\n"
        md += "\n"

    # Eisenhower Matrix
    md += "## Eisenhower Matrix (Urgency vs Importance)\n\n"
    quadrants = {
        "Urgent and Important": [],
        "Not Urgent but Important": [],
        "Urgent but Not Important": [],
        "Not Urgent and Not Important": []
    }

    for task in tasks:
        urgent = task.urgency >= 50
        important = task.importance >= 50
        if urgent and important:
            quadrants["Urgent and Important"].append(task)
        elif not urgent and important:
            quadrants["Not Urgent but Important"].append(task)
        elif urgent and not important:
            quadrants["Urgent but Not Important"].append(task)
        else:
            quadrants["Not Urgent and Not Important"].append(task)

    for quadrant, qtasks in quadrants.items():
        md += f"### {quadrant}\n\n"
        if qtasks:
            md += "| Task ID | Title | Urgency | Importance | Status |\n"
            md += "|---------|-------|---------|------------|--------|\n"
            for task in qtasks:
                md += f"| {task.id} | {task.title} | {task.urgency:.2f} | {task.importance:.2f} | {task.status} |\n"
        else:
            md += "- None\n"
        md += "\n"

    md += "\n## Notes\n\n"
    md += "- Update this dashboard regularly to reflect progress.\n"
    md += "- Use commit messages to update task statuses in \"project_management_state.txt\".\n"
    md += "- This dashboard provides a real-time overview of project progress and workflow status.\n\n"

    md += "## Test Coverage Summary\n\n"
    md += "| Module/Feature | Status |\n"
    md += "|---------------|--------|\n"
    md += "| Task Management Module | Mostly Complete |\n"
    md += "| GitHub Integration Module | Partial |\n"
    md += "| Progress Reporting | Pending |\n"
    md += "| Cross-project Compatibility | Pending |\n\n"
    md += "*See TEST_COVERAGE.txt for detailed test coverage tracking.*\n"

    with open(os.path.join('docs', 'project_management', 'progress_dashboard.md'), 'w', encoding='utf-8') as f:
        f.write(md)
    print("Progress dashboard report updated at docs/project_management/progress_dashboard.md")

def generate_importance_urgency_report(tm: TaskManagement):
    tm.calculate_urgency_importance()
    tasks = tm.prioritize_tasks()

    # Mapping for workflow position and files involved by task title
    workflow_positions = {
        "Develop Project Management Tool": "Initial phase, project setup and architecture design.",
        "Develop Project Management Tool - Subtask Level 1.1": "Early development, security and access control.",
        "Develop Project Management Tool - Subtask Level 1.3": "Mid development, data visualization and reporting.",
        "Develop Project Management Tool - Subtask Level 1.2": "Core functionality development.",
        "Develop Project Management Tool - Subtask Level 2.1.2": "Integration phase.",
        "Develop Project Management Tool - Subtask Level 2.3.2": "Performance tuning and optimization.",
        "Develop Project Management Tool - Subtask Level 2.1.1": "DevOps and automation.",
        "Develop Project Management Tool - Subtask Level 2.3.1": "User engagement and communication.",
        "Develop Project Management Tool - Subtask Level 2.2.1": "Backend API development.",
        "Develop Project Management Tool - Subtask Level 2.2.2": "Security and permissions.",
    }

    files_involved = {
        "Develop Project Management Tool": ["src/task_management.py", "src/progress_report.py", "src/auto_commit.py"],
        "Develop Project Management Tool - Subtask Level 1.1": ["src/task_management.py"],
        "Develop Project Management Tool - Subtask Level 1.3": ["src/progress_report.py"],
        "Develop Project Management Tool - Subtask Level 1.2": ["src/task_management.py"],
        "Develop Project Management Tool - Subtask Level 2.1.2": ["src/task_management.py"],
        "Develop Project Management Tool - Subtask Level 2.3.2": ["src/task_management.py"],
        "Develop Project Management Tool - Subtask Level 2.1.1": ["src/auto_commit.py"],
        "Develop Project Management Tool - Subtask Level 2.3.1": ["src/task_management.py"],
        "Develop Project Management Tool - Subtask Level 2.2.1": ["src/task_management.py"],
        "Develop Project Management Tool - Subtask Level 2.2.2": ["src/task_management.py"],
    }

    # Filter tasks by testing phase
    not_testing_tasks = [t for t in tasks if t.status.lower() != 'testing']
    testing_tasks = [t for t in tasks if t.status.lower() == 'testing']

    # Sort tasks by importance and urgency separately
    important_not_testing = sorted(not_testing_tasks, key=lambda t: t.importance, reverse=True)[:15]
    urgent_not_testing = sorted(not_testing_tasks, key=lambda t: t.urgency, reverse=True)[:15]
    important_testing = sorted(testing_tasks, key=lambda t: t.importance, reverse=True)[:10]
    urgent_testing = sorted(testing_tasks, key=lambda t: t.urgency, reverse=True)[:10]

    def format_task(task):
        wp = workflow_positions.get(task.title, "No workflow position available.")
        fi = files_involved.get(task.title, ["No files listed."])
        fi_str = ", ".join(fi)
        return (f"- **{task.title}** (ID: {task.id})\n"
                f"  - Importance: {task.importance:.2f}\n"
                f"  - Urgency: {task.urgency:.2f}\n"
                f"  - Status: {task.status}\n"
                f"  - Description: {task.description if hasattr(task, 'description') else 'No description available.'}\n"
                f"  - Workflow Position: {wp}\n"
                f"  - Files Involved: {fi_str}\n")

    report = f"# Importance and Urgency Report - {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
    report += "## Top 15 Important Tasks Not in Testing Phase\n"
    report += "\n".join(format_task(t) for t in important_not_testing) if important_not_testing else "- None"
    report += "\n\n## Top 15 Urgent Tasks Not in Testing Phase\n"
    report += "\n".join(format_task(t) for t in urgent_not_testing) if urgent_not_testing else "- None"
    report += "\n\n## Top 10 Important Tasks In Testing Phase\n"
    report += "\n".join(format_task(t) for t in important_testing) if important_testing else "- None"
    report += "\n\n## Top 10 Urgent Tasks In Testing Phase\n"
    report += "\n".join(format_task(t) for t in urgent_testing) if urgent_testing else "- None"
    report += "\n\n*This report is generated automatically and lists the top important and urgent tasks with details and current workflow status, separated by testing phase.*\n"

    with open(IMPORTANCE_URGENCY_REPORT_PATH, 'w', encoding='utf-8') as f:
        f.write(report)
    print(f"Importance and Urgency report updated at {IMPORTANCE_URGENCY_REPORT_PATH}")

if __name__ == "__main__":
    import openpyxl
    from openpyxl.styles import Font, Alignment

    def generate_comprehensive_excel_report(tm: TaskManagement, output_path="docs/project_management/comprehensive_report.xlsx"):
        """
        Generate a comprehensive Excel report with columns:
        Level (لول), Description (شرح), Priority (اولویت), Urgency (فوریت), Status (وضعیت), Progress Percentage (درصد پیشرفت)
        """
        tasks = tm.prioritize_tasks()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Comprehensive Report"

        headers = ["Level", "Description", "Priority", "Urgency", "Status", "Progress Percentage"]
        ws.append(headers)

        # Set header styles
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for task in tasks:
            # Extract level from task title if possible, e.g. "Subtask Level 1.1" -> "1.1"
            level = "1"
            import re
            match = re.search(r"Level ([\\d\\.]+)", task.title)
            if match:
                level = match.group(1)

            description = getattr(task, "description", "") or ""
            priority = round(task.importance, 2) if hasattr(task, "importance") else ""
            urgency = round(task.urgency, 2) if hasattr(task, "urgency") else ""
            status = task.status if hasattr(task, "status") else ""
            progress = 0
            if hasattr(task, "workflow_progress_percentage"):
                if callable(task.workflow_progress_percentage):
                    progress = task.workflow_progress_percentage()
                else:
                    progress = task.workflow_progress_percentage

            row = [level, description, priority, urgency, status, progress]
            ws.append(row)

        # Adjust column widths
        column_widths = [10, 50, 15, 15, 15, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        wb.save(output_path)
        print(f"Comprehensive Excel report generated at {output_path}")

    tm = TaskManagement()
    # For demonstration, generate WBS from example idea
    tm.generate_wbs_from_idea("Develop Project Management Tool")

    # Mark top 15 important tasks not in testing as completed
    tasks = tm.prioritize_tasks()
    important_not_testing = [t for t in tasks if t.status.lower() != 'testing']
    important_not_testing_sorted = sorted(important_not_testing, key=lambda t: t.importance, reverse=True)[:15]
    for task in important_not_testing_sorted:
        tm.mark_task_completed(task.id)

    generate_report(tm)
    generate_importance_urgency_report(tm)
    generate_progress_dashboard_report(tm)
    generate_comprehensive_excel_report(tm)
